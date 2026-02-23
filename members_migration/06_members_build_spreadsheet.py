#!/usr/bin/env python3
"""
Build members_canonical.xlsx from Stage 2 CSV outputs.

Inputs (under out/members/):
- stage2_members_canonical.csv
- stage2_member_activity.csv
- stage2_members_active.csv
- dedup_report.csv

Output:
- members_canonical.xlsx

Rules:
- Deterministic ordering
- Restricted / PII columns hidden
- Freeze headers + filters
- Human-review friendly
"""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ---------- helpers ----------

def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def set_header(ws, headers: List[str]) -> None:
    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)


def autosize(ws, max_width: int = 60) -> None:
    widths = {}
    for row in ws.iter_rows(values_only=True):
        for i, v in enumerate(row, start=1):
            if v is None:
                continue
            widths[i] = max(widths.get(i, 10), len(str(v)) + 2)
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = min(w, max_width)


def freeze(ws, row: int = 2, col: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=row, column=col).coordinate


def hide_columns(ws, headers: List[str], names: List[str]) -> None:
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for n in names:
        if n in idx:
            ws.column_dimensions[get_column_letter(idx[n])].hidden = True


def add_rows(ws, headers: List[str], rows: List[Dict[str, str]]) -> None:
    for r in rows:
        ws.append([r.get(h, "") for h in headers])


# ---------- main ----------

def main() -> None:
    out_dir = Path("out/members")

    members = read_csv(out_dir / "stage2_members_canonical.csv")
    activity = read_csv(out_dir / "stage2_member_activity.csv")
    active = read_csv(out_dir / "stage2_members_active.csv")
    dedup = read_csv(out_dir / "dedup_report.csv") if (out_dir / "dedup_report.csv").exists() else []

    wb = Workbook()
    wb.remove(wb.active)

    # README
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "Members Canonical Dataset (Mirror-Derived)"
    ws["A1"].font = Font(bold=True, size=14)

    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    readme_lines = [
        "",
        "Source: Local public mirror of footbag.org",
        "Purpose: Test + migration dataset for modernization project",
        "",
        "ACTIVE definition:",
        "- Event-linked role evidence OR last-login presence",
        "",
        "Privacy:",
        "- restricted_* columns are hidden by default",
        "",
        f"Generated (UTC): {ts}",
    ]
    for i, line in enumerate(readme_lines, start=3):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 110

    # Members
    ws = wb.create_sheet("Members")
    members_headers = list(members[0].keys())
    set_header(ws, members_headers)
    add_rows(ws, members_headers, members)
    freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(members_headers))}{len(members)+1}"
    hide_columns(ws, members_headers, [
        "restricted_last_login_raw",
        "restricted_pii_email_raw",
        "restricted_pii_phone_raw",
        "restricted_pii_address_raw",
    ])
    autosize(ws)

    # ActivityEvidence
    ws = wb.create_sheet("ActivityEvidence")
    act_headers = list(activity[0].keys())
    set_header(ws, act_headers)
    add_rows(ws, act_headers, activity)
    freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(act_headers))}{len(activity)+1}"
    autosize(ws)

    # ActiveMembers
    ws = wb.create_sheet("ActiveMembers")
    active_headers = list(active[0].keys())
    set_header(ws, active_headers)
    add_rows(ws, active_headers, active)
    freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(active_headers))}{len(active)+1}"
    autosize(ws)

    # DuplicatesReview
    ws = wb.create_sheet("DuplicatesReview")
    dup_headers = list(dedup[0].keys()) if dedup else [
        "duplicate_key",
        "member_ids",
        "legacy_member_ids",
        "legacy_usernames",
        "reason",
        "recommended_action",
    ]
    set_header(ws, dup_headers)
    add_rows(ws, dup_headers, dedup)
    freeze(ws)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(dup_headers))}{len(dedup)+1}"
    autosize(ws)

    # FieldCoverage
    ws = wb.create_sheet("FieldCoverage")
    headers = ["field", "filled", "total", "pct"]
    set_header(ws, headers)

    total = len(members)
    fields = [
        "legacy_username",
        "name_display",
        "joined_raw",
        "club_ids",
        "club_names",
        "active",
        "active_confidence",
        "evidence_count",
    ]
    for f in fields:
        filled = sum(1 for r in members if (r.get(f) or "").strip())
        pct = (filled / total * 100.0) if total else 0.0
        ws.append([f, filled, total, f"{pct:.1f}%"])

    freeze(ws)
    autosize(ws)

    out_path = out_dir / "members_canonical.xlsx"
    wb.save(out_path)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
